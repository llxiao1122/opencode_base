"""
送站补单表生成脚本
结构：补单表/收货人名/合同号/PO_厂验清单.xlsx + PO_送站签收单.xlsx
每个 xlsx 只包含一个采购订单号的数据，只保留对应 sheet。
只生成 14号线一期 数据。
"""

import os, shutil, sys, traceback
from collections import OrderedDict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_FILE = os.path.join(SCRIPT_DIR, "送站待补单第三批-20260729.xlsx")
OUT_DIR = os.path.join(SCRIPT_DIR, "补单表")

FIXED_ADDRESS = "郑州市中原区14号线铁炉西车辆段"
THIN_BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
DATA_FONT = Font(name="微软雅黑", size=9)
DATA_ALIGN = Alignment(wrap_text=True, vertical="center")
LINE_FILTER = "运营14号线一期库存"


def safe_name(s):
    for ch in r'\/:*?"<>|':
        s = s.replace(ch, "_")
    return s.strip() or "未知"


def detect_org_label(year):
    return "物资管理中心" if year == 2026 else "物资管理部"


def extract_year(contract_no):
    try:
        parts = str(contract_no).split("-")
        if len(parts) >= 3:
            code = parts[2]
            if code and code[0].isalpha():
                return 2000 + int(code[1:3])
    except:
        pass
    return None


def build_groups(ws1):
    """返回 {(recipient, contract_no, po): [row_numbers]}"""
    groups = OrderedDict()
    for r in range(2, ws1.max_row + 1):
        line = str(ws1.cell(r, 6).value or "").strip()
        if line != LINE_FILTER:
            continue
        po = ws1.cell(r, 2).value
        if not po:
            continue
        recipient = str(ws1.cell(r, 21).value or "未知").strip()
        contract_no = str(ws1.cell(r, 9).value or "").strip()
        po = str(po).strip()
        key = (recipient, contract_no, po)
        groups.setdefault(key, []).append(r)
    return groups


def make_clean_template(sheet_name):
    wb = load_workbook(SRC_FILE)
    for sn in list(wb.sheetnames):
        if sn != sheet_name:
            del wb[sn]
    ws = wb[sheet_name]
    data_start = 5 if sheet_name == "厂验清单" else 4
    if ws.max_row >= data_start:
        ws.delete_rows(data_start, ws.max_row - data_start + 1)
    return wb, ws


def write_data_cv(ws, rows, ws1):
    rn = 5
    for sr in rows:
        vals = [str(ws1.cell(sr, c).value or "") for c in [2, 15, 6, 23, 17, 18, 19]]
        vals.append(ws1.cell(sr, 20).value or 0)
        vals += [None, None, None]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(rn, ci, v)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
        rn += 1


def write_data_qs(ws, rows, ws1):
    rn = 4
    for sr in rows:
        vals = [str(ws1.cell(sr, c).value or "") for c in [2, 15]]
        vals += [None]
        vals += [str(ws1.cell(sr, c).value or "") for c in [17, 18, 19]]
        vals.append(ws1.cell(sr, 20).value or 0)
        vals += [str(ws1.cell(sr, c).value or "") for c in [6, 23, 22, 21, 24]]
        vals += [None, None, None, None]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(rn, ci, v)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
        rn += 1


def main():
    print("正在读取源文件...")
    wb_src = load_workbook(SRC_FILE, data_only=True)
    ws1 = wb_src["汇总清单"]

    groups = build_groups(ws1)
    total = len(groups)
    print(f"共 {total} 张补单（14号线一期）")

    if os.path.exists(OUT_DIR):
        shutil.rmtree(OUT_DIR)

    seen_rec = set()
    for idx, ((recipient, contract_no, po), rows) in enumerate(groups.items(), 1):
        rec_folder = safe_name(recipient)
        ct_folder = safe_name(contract_no) if contract_no else "无合同号"
        folder_path = os.path.join(OUT_DIR, rec_folder, ct_folder)
        os.makedirs(folder_path, exist_ok=True)

        r0 = rows[0]
        contract_name = ws1.cell(r0, 10).value or ""
        supplier = ws1.cell(r0, 7).value or ""

        arrival_raw = ws1.cell(r0, 13).value
        if isinstance(arrival_raw, datetime):
            factory_date = arrival_raw
        elif arrival_raw:
            try:
                factory_date = datetime.strptime(str(arrival_raw)[:10], "%Y-%m-%d")
            except:
                factory_date = datetime.now()
        else:
            factory_date = datetime.now()

        year = extract_year(contract_no) if contract_no else None
        org_label = detect_org_label(year) if year else "物资管理部"

        po_safe = safe_name(po)

        # 厂验清单
        wb_cv, ws_cv = make_clean_template("厂验清单")
        ws_cv.cell(2, 2, contract_name)
        ws_cv.cell(2, 8, contract_no)
        ws_cv.cell(2, 11, f"供应商：{supplier}")
        ws_cv.cell(3, 2, FIXED_ADDRESS)
        ws_cv.cell(3, 8, factory_date.strftime("%Y-%m-%d"))
        ws_cv.cell(3, 11, f"{org_label}验收人员（姓名+工号）：")
        write_data_cv(ws_cv, rows, ws1)
        wb_cv.save(os.path.join(folder_path, f"{po_safe}_厂验清单.xlsx"))
        wb_cv.close()

        # 送站签收单
        wb_qs, ws_qs = make_clean_template("送站签收单")
        ws_qs.cell(2, 3, supplier)
        ws_qs.cell(2, 6, contract_name)
        ws_qs.cell(2, 14, contract_no)
        write_data_qs(ws_qs, rows, ws1)
        wb_qs.save(os.path.join(folder_path, f"{po_safe}_送站签收单.xlsx"))
        wb_qs.close()

        if recipient not in seen_rec:
            seen_rec.add(recipient)
            print(f"[{recipient}]")

    print(f"\n完成：{total} 张补单，输出至 {OUT_DIR}/")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n错误：{e}", file=sys.stderr)
        traceback.print_exc()
    finally:
        input("\n按 Enter 键退出...")
